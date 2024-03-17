import pandas as pd
from openpyxl import Workbook
import os
import xlsxwriter as xls
import xlwt

def write2(fileName, df):
    path = getPath(fileName)
    wb = xls.Workbook(path)
    ws = wb.add_worksheet("Sheet1")
    ws.write(0, 0, "#")
    ws.write(0, 1, "Datum")
    ws.write(0, 2, "Tid")
    ws.write(0, 3, "Distrikt")
    ws.write(0, 4, "Tjänst")
    ws.write(0, 5, "Kostnad")
    ws.write(0, 6, "Resor (km)")
    ws.write(0, 7, "Övrigt")

    for index, entry in enumerate(df):
        ws.write(index+1, 0, df[0])
        ws.write(index+1, 1, df[1])
        ws.write(index+1, 2, df[2])
        ws.write(index+1, 3, df[3])
        ws.write(index+1, 4, df[4])
        ws.write(index+1, 5, df[5])
        ws.write(index+1, 6, df[6])
        ws.write(index+1, 7, "")

def getPath(fileName):
    path = '/Users/victorpekkari/Documents/salg/outputFiles/'
    #path = '/outputFiles/'
    format = '.xls'
    file = path + fileName + format
    os.makedirs(os.path.dirname(file), exist_ok=True)
    return file

def write(fileName, df):
    path = '/Users/victorpekkari/Documents/salg/outputFiles/'
    format = '.xls'
    file = path + fileName + format
    os.makedirs(os.path.dirname(file), exist_ok=True)

    if not os.path.exists(file):
        create_xls_file(file)

    # Create a new workbook and sheet
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet2')

    # Write column names to the first row
    for j, col_name in enumerate(df.columns):
        sheet.write(0, j, col_name)

    # Write the DataFrame data
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            sheet.write(i+1, j, value)  # Start writing data from the second row

    # Save the workbook to the file
    workbook.save(file)

def create_xls_file(file_path):
    # Create an empty file
    open(file_path, 'w').close()

def delete_contents(file):
    wb = Workbook()
    ws = wb.active
    ws = wb.create_sheet("Sheet1")    
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    
    # Save the changes to the file
    wb.save(file)
