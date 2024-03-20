from openpyxl import Workbook
import os
import xlsxwriter as xlsx
from Constants import *
import datetime

def getPath(fileName):
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    format = '.xls'
    file = desktop + path + fileName + format
    os.makedirs(os.path.dirname(file), exist_ok=True)
    return file

def format_number(number_str):
    number = int(number_str)
    formatted_number = '{:,.0f}'.format(number).replace(",", " ")
    return str(formatted_number)

def getOccurences(df, column_name, value):
    map = {}
    for val in value:
        map[val] = (df[column_name] == val).sum()
    return map

def write(fileName, df, place):
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    file = desktop + path + fileName + file_format
    os.makedirs(os.path.dirname(file), exist_ok=True)

    if not os.path.exists(file):
        create_xls_file(file)

    # Create a new workbook and sheet
    workbook = xlsx.Workbook(file)
    workbook = xlsx.Workbook(file, {'nan_inf_to_errors': True})
    sheet = workbook.add_worksheet('Sheet2')
    yellow_format = workbook.add_format({'bg_color': 'yellow', 'font_size': 14, 'bold':True, 'border': 1})

    column_width = 20  # 20 characters wide
    sheet.set_column(0, len(df.columns) - 1, column_width)
    header_format = workbook.add_format({'font_size': 20})  # Adjust font size as needed
    small_text = workbook.add_format({'font_size': 10})
    light_blue = workbook.add_format({'bg_color': '#ADD8E6', 'border': 1})
    header_light_blue = workbook.add_format({'bg_color': '#ADD8E6', 'border': 1, 'bold':True})
    black_borders = workbook.add_format({'border': 1})

    sheet.write(0, 0, fileName.split("/")[0], header_format)
    sheet.write(0, 1, 'skapades: ' + str(datetime.date.today()), small_text)

    # Write column names to the first row
    for j, col_name in enumerate(df.columns):
        #if j == 0 or j == len(df.columns) - 1:  # First and last columns    
        sheet.write(startWrite-1, j, col_name, yellow_format)

    map = getOccurences(df, 'Tjänst', [value for value in taskMapping.values()])

    # Write the occurences of the different tasks
    sheet.write(startWrite-3, 0, 'Antal', header_light_blue)
    sheet.write(startWrite-2, 0, 'Pris', header_light_blue)
    
    for j, task in enumerate(map.keys()):
        sheet.write(startWrite-4, 1 + j, task, header_light_blue)
        sheet.write(startWrite-3, 1 + j, map[task], light_blue)
        price = price_place_task[place][task]
        if str(price).isdigit():
            sheet.write(startWrite-2, 1 + j, format_number(str(int(price) * int(map[task]))), light_blue)
        else:
            sheet.write(startWrite-2, 1 + j, price, light_blue)

    # Write the DataFrame data
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            sheet.write(startWrite + i, j, value)  # Start writing data from the third row

    # Save the workbook to the file
    workbook.close()


def create_xls_file(file_path):
    # Create an empty file
    open(file_path, 'w').close()

def delete_contents(file):
    wb = Workbook()
    ws = wb.active
    ws = wb.create_sheet("Sheet1")    
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    
    # Save the changes to the file
    wb.save(file)
